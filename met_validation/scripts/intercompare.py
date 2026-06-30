"""Intercomparison: measured met data vs WRFLakes-CALMET vs ERA5-CALMET.

For each of the six Qatar monitoring stations:
  1. Read hourly observations from the Excel workbook (one sheet per station).
  2. Read CALMET-Lite "Timeseries at gridpoint" exports for the two CALMET
     runs (WRFLakes-driven and ERA5-driven).
  3. Align all three series in local time (Asia/Qatar = UTC+3).  WRFLakes
     CALMET was run in local time so its stamps are used as-is.  The ERA5
     CALMET run was in UTC, so we shift its stamps +3 h.  Observations are
     station time, which we treat as local.
  4. Restrict to 2022-07-01 .. 2022-09-30 (extended to whatever ERA5 covers).
  5. Compute scalar statistics for temperature, wind speed; circular
     statistics for wind direction; emit per-station/per-variable plots
     and a combined CSV + markdown report.

Run from the repo root:

    python met_validation/scripts/intercompare.py

Outputs go under ``met_validation/output/`` and the script is idempotent.
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd


_LOG = logging.getLogger("intercompare")


# ---------------------------------------------------------------------------
# Station catalog
# ---------------------------------------------------------------------------
#
# Coordinates are UTM Zone 39N metres (Location X/Y from the GIS map).  The
# CALMET filenames carry the same coordinates in kilometres, so we match by
# the km value with a small tolerance to absorb rounding (e.g. one file is
# X556.224 and another X556.226 for the same station).


@dataclass(frozen=True)
class Station:
    name: str          # human-readable name
    sheet: str         # exact Excel sheet name
    x_km: float
    y_km: float


STATIONS: tuple[Station, ...] = (
    Station("Al Jumaliya", "JUMALIYA ",   507.842, 2834.220),
    Station("Al Shamal",   "Shamal",      521.152, 2889.652),
    Station("Al Dhakira",  "DHAKIRA ",    550.368, 2839.546),
    Station("Al Khor",     "Al Khor",     553.319, 2845.125),
    Station("RLIC Camp",   "RLIC Camp",   556.223, 2859.432),
    Station("RLIC Port",   "NEW PORT ",   559.135, 2867.296),
)


# Round-half-up to the nearest hour (so 21:59 -> 22:00; 21:30 -> 22:00).
# pandas' .round('H') uses banker's rounding, which would send 21:30 -> 22:00
# but 20:30 -> 20:00 (round-half-to-even).  We want a consistent behaviour.
def _round_half_up_to_hour(ts: pd.Series) -> pd.Series:
    return (ts + pd.Timedelta(minutes=30)).dt.floor("h")


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------


_DAT_HEADER_LINES = 4  # 4 leading non-data lines in the CALMET-Lite export


def read_calmet_dat(path: Path) -> pd.DataFrame:
    """Parse a CALMET-Lite "Timeseries at gridpoint" .dat file.

    Returns a DataFrame indexed by the start-time DatetimeIndex with
    columns ``ws``, ``wd``, ``t_k`` (and ``t_c`` for convenience).
    """
    df = pd.read_csv(
        path, sep=r"\s+", skiprows=_DAT_HEADER_LINES, header=None, engine="python",
    )
    # The CALMET export carries both START (cols 0..4) and END (cols 5..9)
    # of each averaging interval; we index on START.
    if df.shape[1] < 13:
        raise ValueError(
            f"{path.name}: expected >= 13 columns, got {df.shape[1]}"
        )
    start = pd.to_datetime({
        "year": df[0], "month": df[1], "day": df[2],
        "hour": df[3], "minute": 0, "second": df[4],
    })
    out = pd.DataFrame({
        "ws":  df[10].astype(float).to_numpy(),
        "wd":  df[11].astype(float).to_numpy(),
        "t_k": df[12].astype(float).to_numpy(),
    })
    out.index = pd.DatetimeIndex(start.to_numpy(), name="time")
    out["t_c"] = out["t_k"] - 273.15
    return out


# Two header layouts exist in this workbook:
#   layout A (Shamal, Al Khor, RLIC Camp): headers on row 3 (1-based)
#   layout B (NEW PORT, DHAKIRA, JUMALIYA): headers on row 2 (1-based)
# We auto-detect by searching the first 5 rows for one containing a cell
# whose text matches /^Date/.  Missing-data tokens like "CC" / "SM" are
# treated as NaN.
_NA_TOKENS = {"CC", "SM", "NA", "N/A", "--"}


def read_obs(xlsx: Path, sheet: str) -> pd.DataFrame:
    """Read a single station sheet from the workbook.

    Returns a DataFrame with index = obs hourly timestamps in local time
    and columns ``ws`` (m/s), ``wd`` (deg), ``t_c`` (°C).  Invalid /
    missing cells become NaN.
    """
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"sheet {sheet!r} not in workbook ({wb.sheetnames!r})")
    ws = wb[sheet]

    # Scan first 5 rows for the header row — identified by a leading cell
    # whose text starts with "Date".
    header_row: list[str] = []
    header_row_idx = -1
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ridx > 5:
            break
        cells = [str(c).strip() if c is not None else "" for c in row]
        if cells and cells[0].lower().startswith("date"):
            header_row = cells
            header_row_idx = ridx
            break
    if header_row_idx < 0:
        raise ValueError(f"{sheet!r}: header row not found in first 5 rows")

    def _col(label: str) -> int:
        for i, h in enumerate(header_row):
            if h.upper().strip() == label.upper():
                return i
        raise KeyError(
            f"{sheet!r}: header column {label!r} not in {header_row!r}"
        )

    i_time = 0                       # always first column ("Date & Time"/"Date / Time")
    i_ws = _col("WS")
    i_wd = _col("WD")
    i_t = _col("TEMP")

    times: list = []
    ws_vals: list = []
    wd_vals: list = []
    t_vals: list = []

    # The "units" row is right after headers; data starts at header_row+2
    # for layouts that have a blank separator row, or header_row+1 in
    # layouts without.  We just keep reading and skip rows whose first
    # column isn't parseable as a timestamp.
    wb.close()
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[sheet]
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ridx <= header_row_idx + 1:
            continue
        t = row[i_time]
        if t is None:
            continue
        times.append(t)
        ws_vals.append(_to_float(row[i_ws]))
        wd_vals.append(_to_float(row[i_wd]))
        t_vals.append(_to_float(row[i_t]))
    wb.close()

    idx = pd.to_datetime(pd.Series(times), errors="coerce")
    df = pd.DataFrame(
        {"ws": ws_vals, "wd": wd_vals, "t_c": t_vals},
        index=pd.DatetimeIndex(idx, name="time"),
    )
    df = df[~df.index.isna()]
    df.index = _round_half_up_to_hour(pd.Series(df.index)).to_numpy()
    df.index.name = "time"
    df = df[~df.index.duplicated(keep="first")]
    df = df.sort_index()
    return df


def _to_float(v) -> float:
    if v is None:
        return float("nan")
    if isinstance(v, str):
        s = v.strip()
        if not s or s.upper() in _NA_TOKENS:
            return float("nan")
        try:
            return float(s)
        except ValueError:
            return float("nan")
    try:
        return float(v)
    except (TypeError, ValueError):
        return float("nan")


# ---------------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------------


def _circular_diff(model: np.ndarray, obs: np.ndarray) -> np.ndarray:
    """Signed wind-direction error wrapped to (-180, 180]."""
    d = (model - obs + 180.0) % 360.0 - 180.0
    return d


def _stats_scalar(obs: pd.Series, mod: pd.Series) -> dict[str, float]:
    """bias / MAE / RMSE / R / N / IOA / FB."""
    df = pd.concat([obs.rename("o"), mod.rename("m")], axis=1).dropna()
    n = len(df)
    if n < 2:
        return {k: float("nan") for k in
                ("n", "bias", "mae", "rmse", "r", "ioa", "fb",
                 "obs_mean", "mod_mean")}
    o = df["o"].to_numpy(dtype=float)
    m = df["m"].to_numpy(dtype=float)
    diff = m - o
    bias = float(np.mean(diff))
    mae = float(np.mean(np.abs(diff)))
    rmse = float(np.sqrt(np.mean(diff**2)))
    o_mean = float(np.mean(o))
    m_mean = float(np.mean(m))
    if np.std(o) > 0 and np.std(m) > 0:
        r = float(np.corrcoef(o, m)[0, 1])
    else:
        r = float("nan")
    denom = np.sum((np.abs(m - o_mean) + np.abs(o - o_mean))**2)
    ioa = 1.0 - float(np.sum(diff**2) / denom) if denom > 0 else float("nan")
    fb_denom = (m_mean + o_mean) / 2.0
    fb = float(bias / fb_denom) if fb_denom != 0 else float("nan")
    return {"n": n, "bias": bias, "mae": mae, "rmse": rmse, "r": r,
            "ioa": ioa, "fb": fb,
            "obs_mean": o_mean, "mod_mean": m_mean}


def _stats_wind_dir(
    obs_wd: pd.Series, mod_wd: pd.Series,
    obs_ws: pd.Series | None = None, mod_ws: pd.Series | None = None,
) -> dict[str, float]:
    """Circular WD stats + (optional) vector RMSE if WS supplied.

    Vector RMSE is the RMS of |(u_m,v_m) - (u_o,v_o)|, i.e. the wind
    vector error treating model and obs as 2-D vectors.
    """
    df = pd.concat(
        [obs_wd.rename("o"), mod_wd.rename("m")], axis=1
    ).dropna()
    n = len(df)
    if n < 2:
        base = {k: float("nan") for k in
                ("n", "bias", "mae", "rmse", "ioa",
                 "within_22_5", "within_45")}
        if obs_ws is not None:
            base["vector_rmse"] = float("nan")
        return base
    o = df["o"].to_numpy(dtype=float)
    m = df["m"].to_numpy(dtype=float)
    diff = _circular_diff(m, o)
    bias = float(np.mean(diff))
    mae = float(np.mean(np.abs(diff)))
    rmse = float(np.sqrt(np.mean(diff**2)))
    within_225 = float(np.mean(np.abs(diff) <= 22.5))
    within_45 = float(np.mean(np.abs(diff) <= 45.0))
    # IOA on a circular variable is not standard; report scalar IOA on
    # the signed circular difference for completeness.
    o_mean = 0.0    # circular reference; mean direction not meaningful
    denom = float(np.sum((np.abs(m - o_mean) + np.abs(o - o_mean))**2))
    ioa = (1.0 - float(np.sum(diff**2)) / denom) if denom > 0 else float("nan")
    out: dict[str, float] = {
        "n": n, "bias": bias, "mae": mae, "rmse": rmse, "ioa": ioa,
        "within_22_5": within_225, "within_45": within_45,
    }
    if obs_ws is not None and mod_ws is not None:
        joined = pd.concat([
            obs_wd.rename("od"), mod_wd.rename("md"),
            obs_ws.rename("os"), mod_ws.rename("ms"),
        ], axis=1).dropna()
        if len(joined) >= 2:
            od = np.deg2rad(joined["od"].to_numpy())
            md = np.deg2rad(joined["md"].to_numpy())
            os_ = joined["os"].to_numpy()
            ms_ = joined["ms"].to_numpy()
            # Met convention: wind direction is the direction the wind
            # comes FROM.  u/v relative to that:
            uo = -os_ * np.sin(od)
            vo = -os_ * np.cos(od)
            um = -ms_ * np.sin(md)
            vm = -ms_ * np.cos(md)
            vrmse = float(np.sqrt(np.mean((um - uo)**2 + (vm - vo)**2)))
            out["vector_rmse"] = vrmse
        else:
            out["vector_rmse"] = float("nan")
    return out


# ---------------------------------------------------------------------------
# Per-station processing
# ---------------------------------------------------------------------------


_FNAME_RE = re.compile(
    r"X(?P<x>[0-9.]+)_Y(?P<y>[0-9.]+).*_(?P<tag>WRFLakes|ERA5)\.dat$"
)


def discover_dat_files(data_dir: Path) -> dict[str, dict[tuple[int, int], Path]]:
    """Group .dat files by (round(x*10), round(y*10)) so 556.224 and 556.226
    land in the same bucket; tag key is 'WRFLakes' or 'ERA5'.
    """
    files: dict[tuple[int, int], dict[str, Path]] = {}
    for p in sorted(data_dir.glob("*.dat")):
        m = _FNAME_RE.search(p.name)
        if not m:
            _LOG.warning("Ignoring file (no X/Y/tag match): %s", p.name)
            continue
        x = float(m.group("x"))
        y = float(m.group("y"))
        key = (round(x * 10), round(y * 10))
        files.setdefault(key, {})[m.group("tag")] = p
    return files


def _match_station(s: Station, files_by_coord: dict) -> dict[str, Path]:
    """Find the dat-file pair for a station, allowing 0.05 km coordinate
    tolerance (the WRFLakes vs ERA5 X disagree by 0.002 km for RLIC Camp).
    """
    best_key = None
    best_dist = float("inf")
    for key in files_by_coord:
        x = key[0] / 10.0
        y = key[1] / 10.0
        d = ((x - s.x_km)**2 + (y - s.y_km)**2)**0.5
        if d < best_dist:
            best_dist = d
            best_key = key
    if best_key is None or best_dist > 0.2:
        return {}
    return files_by_coord[best_key]


# ---------------------------------------------------------------------------
# Plotting
# ---------------------------------------------------------------------------


def _plot_timeseries(
    station: str, var_label: str,
    obs: pd.Series, wrf: pd.Series, era: pd.Series, units: str,
    out: Path,
) -> None:
    fig, ax = plt.subplots(figsize=(11, 4))
    if obs is not None:
        ax.plot(obs.index, obs.values, label="obs", color="black", lw=0.8)
    if wrf is not None:
        ax.plot(wrf.index, wrf.values, label="WRFLakes-CALMET",
                color="tab:blue", lw=0.7, alpha=0.85)
    if era is not None:
        ax.plot(era.index, era.values, label="ERA5-CALMET",
                color="tab:red", lw=0.7, alpha=0.85)
    ax.set_title(f"{station} — {var_label}")
    ax.set_ylabel(f"{var_label} ({units})")
    ax.set_xlabel("local time (UTC+3)")
    ax.legend(loc="upper right", fontsize=8)
    ax.grid(alpha=0.3)
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(out, dpi=120)
    plt.close(fig)


def _plot_scatter(
    station: str, var_label: str,
    obs: pd.Series, mod: pd.Series, label: str, units: str,
    out: Path,
) -> None:
    df = pd.concat([obs.rename("o"), mod.rename("m")], axis=1).dropna()
    if len(df) < 2:
        return
    fig, ax = plt.subplots(figsize=(5, 5))
    ax.scatter(df["o"], df["m"], s=4, alpha=0.4)
    lo = float(min(df["o"].min(), df["m"].min()))
    hi = float(max(df["o"].max(), df["m"].max()))
    ax.plot([lo, hi], [lo, hi], "k--", lw=0.7, alpha=0.7)
    ax.set_xlabel(f"obs {var_label} ({units})")
    ax.set_ylabel(f"{label} {var_label} ({units})")
    ax.set_title(f"{station} — {label} vs obs")
    ax.grid(alpha=0.3)
    ax.set_aspect("equal", adjustable="box")
    fig.tight_layout()
    fig.savefig(out, dpi=120)
    plt.close(fig)


# ---------------------------------------------------------------------------
# Top-level driver
# ---------------------------------------------------------------------------


@dataclass
class StationResult:
    station: Station
    stats_rows: list[dict]


def _align_to_local(
    obs: pd.DataFrame, wrf: pd.DataFrame, era: pd.DataFrame,
    start: pd.Timestamp, end: pd.Timestamp,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    # Observations: station local time, already hourly (rounded).
    # WRFLakes-CALMET: run was UTC+3, stamps already local — pass through.
    # ERA5-CALMET: run was UTC, shift +3 h to get local stamps.
    era_local = era.copy()
    era_local.index = era_local.index + pd.Timedelta(hours=3)

    def clip(df):
        return df[(df.index >= start) & (df.index <= end)]

    return clip(obs), clip(wrf), clip(era_local)


def _evaluate_station(
    s: Station,
    obs_xlsx: Path,
    files: dict[str, Path],
    start: pd.Timestamp,
    end: pd.Timestamp,
    out_dir: Path,
) -> list[dict]:
    if not {"WRFLakes", "ERA5"}.issubset(files):
        _LOG.error("Station %s missing model files (%s); skipping",
                   s.name, list(files))
        return []

    _LOG.info("Station %s: reading observations from sheet %r",
              s.name, s.sheet)
    obs = read_obs(obs_xlsx, s.sheet)
    _LOG.info("Station %s: reading WRFLakes %s", s.name, files["WRFLakes"].name)
    wrf = read_calmet_dat(files["WRFLakes"])
    _LOG.info("Station %s: reading ERA5     %s", s.name, files["ERA5"].name)
    era = read_calmet_dat(files["ERA5"])

    obs, wrf, era = _align_to_local(obs, wrf, era, start, end)
    _LOG.info(
        "  joint window: %d obs hrs, %d WRFLakes hrs, %d ERA5 hrs",
        len(obs), len(wrf), len(era),
    )

    rows: list[dict] = []

    # ----- Temperature (°C) -----
    for tag, mod in (("WRFLakes", wrf["t_c"]), ("ERA5", era["t_c"])):
        st = _stats_scalar(obs["t_c"], mod)
        st.update(dict(station=s.name, variable="T (degC)", model=tag))
        rows.append(st)
    _plot_timeseries(
        s.name, "Temperature", obs["t_c"], wrf["t_c"], era["t_c"], "°C",
        out_dir / f"{_slug(s.name)}_timeseries_T.png",
    )
    _plot_scatter(s.name, "Temperature", obs["t_c"], wrf["t_c"],
                  "WRFLakes-CALMET", "°C",
                  out_dir / f"{_slug(s.name)}_scatter_T_WRFLakes.png")
    _plot_scatter(s.name, "Temperature", obs["t_c"], era["t_c"],
                  "ERA5-CALMET", "°C",
                  out_dir / f"{_slug(s.name)}_scatter_T_ERA5.png")

    # ----- Wind speed (m/s) -----
    for tag, mod in (("WRFLakes", wrf["ws"]), ("ERA5", era["ws"])):
        st = _stats_scalar(obs["ws"], mod)
        st.update(dict(station=s.name, variable="WS (m/s)", model=tag))
        rows.append(st)
    _plot_timeseries(
        s.name, "Wind speed", obs["ws"], wrf["ws"], era["ws"], "m/s",
        out_dir / f"{_slug(s.name)}_timeseries_WS.png",
    )
    _plot_scatter(s.name, "Wind speed", obs["ws"], wrf["ws"],
                  "WRFLakes-CALMET", "m/s",
                  out_dir / f"{_slug(s.name)}_scatter_WS_WRFLakes.png")
    _plot_scatter(s.name, "Wind speed", obs["ws"], era["ws"],
                  "ERA5-CALMET", "m/s",
                  out_dir / f"{_slug(s.name)}_scatter_WS_ERA5.png")

    # ----- Wind direction (deg) — circular -----
    for tag, mod_wd, mod_ws in (
        ("WRFLakes", wrf["wd"], wrf["ws"]),
        ("ERA5",     era["wd"], era["ws"]),
    ):
        st = _stats_wind_dir(obs["wd"], mod_wd, obs["ws"], mod_ws)
        st.update(dict(station=s.name, variable="WD (deg)", model=tag))
        rows.append(st)
    _plot_timeseries(
        s.name, "Wind direction", obs["wd"], wrf["wd"], era["wd"], "deg",
        out_dir / f"{_slug(s.name)}_timeseries_WD.png",
    )
    # Skip WD scatter — meaningless without unwrap; the time series + the
    # within-22.5/45 deg metrics carry the comparison.

    return rows


def _slug(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")


def _df_to_markdown(df: pd.DataFrame) -> str:
    """Render a DataFrame as a GitHub-flavoured markdown table without
    needing the optional ``tabulate`` dependency.
    """
    cols = list(df.columns)
    rows = [[str(v) for v in r] for r in df.itertuples(index=False, name=None)]
    widths = [
        max(len(cols[i]), max((len(r[i]) for r in rows), default=0))
        for i in range(len(cols))
    ]
    def fmt(row: list[str]) -> str:
        return "| " + " | ".join(c.ljust(w) for c, w in zip(row, widths)) + " |"
    sep = "| " + " | ".join("-" * w for w in widths) + " |"
    return "\n".join([fmt(cols), sep] + [fmt(r) for r in rows])


def _write_report(
    df: pd.DataFrame, out_dir: Path, start: pd.Timestamp, end: pd.Timestamp,
) -> None:
    md = out_dir / "report.md"
    lines: list[str] = []
    lines.append("# Met-station intercomparison\n")
    lines.append(
        f"Window: **{start:%Y-%m-%d %H:%M}** to **{end:%Y-%m-%d %H:%M}** "
        f"local time (Asia/Qatar, UTC+3)\n"
    )
    lines.append(
        "All statistics computed on the inner-join of obs / WRFLakes-CALMET "
        "/ ERA5-CALMET hourly time series after rounding the observation "
        "timestamps to the nearest hour (half-up).\n"
    )
    lines.append("Generated by `met_validation/scripts/intercompare.py`.\n")

    # Flag stations whose observation series is entirely empty in the
    # joint window so the empty rows in the per-variable tables below
    # make sense to the reader.
    empty = (
        df.assign(has=df["n"].notna() & (df["n"] > 0))
          .groupby("station")["has"].any()
    )
    empty_stations = sorted([s for s, v in empty.items() if not v])
    if empty_stations:
        lines.append(
            f"\n> **Note:** the following stations have no observed met "
            f"data in the joint window and appear as empty rows: "
            f"**{', '.join(empty_stations)}**.\n"
        )

    for var in ("T (degC)", "WS (m/s)", "WD (deg)"):
        lines.append(f"\n## {var}\n")
        sub = df[df["variable"] == var].copy()
        if sub.empty:
            continue
        cols = ["station", "model", "n", "bias", "mae", "rmse",
                "r", "ioa", "fb"]
        if var.startswith("WD"):
            cols = ["station", "model", "n", "bias", "mae", "rmse", "ioa",
                    "within_22_5", "within_45", "vector_rmse"]
        sub = sub[[c for c in cols if c in sub.columns]]
        # Format numeric columns to 3 sig figs.
        sub_disp = sub.copy()
        for c in sub_disp.columns:
            if pd.api.types.is_numeric_dtype(sub_disp[c]):
                sub_disp[c] = sub_disp[c].map(
                    lambda v: "" if pd.isna(v)
                    else (f"{int(v)}" if c == "n" else f"{v:.3g}")
                )
        lines.append(_df_to_markdown(sub_disp) + "\n")

    # Cross-model picks: by station and variable, list which model has
    # the lower RMSE.  Quick "who wins where" summary for the reader.
    lines.append("\n## Model selection by lowest RMSE\n")
    pick_rows: list[list[str]] = []
    for var in ("T (degC)", "WS (m/s)", "WD (deg)"):
        sub = df[df["variable"] == var].dropna(subset=["rmse"])
        if sub.empty:
            continue
        for station, grp in sub.groupby("station"):
            best = grp.loc[grp["rmse"].idxmin()]
            other = grp.loc[grp["rmse"].idxmax()]
            pick_rows.append([
                station, var, str(best["model"]),
                f"{best['rmse']:.3g}", f"{other['rmse']:.3g}",
                f"{other['rmse'] - best['rmse']:+.3g}",
            ])
    if pick_rows:
        pick_df = pd.DataFrame(
            pick_rows,
            columns=["station", "variable", "lower_RMSE_model",
                     "best_rmse", "other_rmse", "delta"],
        )
        lines.append(_df_to_markdown(pick_df) + "\n")

    # Plot index.
    out_dir = md.parent
    pngs = sorted(out_dir.glob("*.png"))
    if pngs:
        lines.append("\n## Plots\n")
        for p in pngs:
            lines.append(f"- [`{p.name}`]({p.name})")
        lines.append("")

    md.write_text("\n".join(lines), encoding="utf-8")
    _LOG.info("Wrote %s", md)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description="Intercompare obs vs WRFLakes-CALMET vs ERA5-CALMET"
    )
    ap.add_argument("--data-dir", type=Path,
                    default=Path("met_validation"))
    ap.add_argument("--xlsx", type=Path,
                    default=Path("met_validation/Mess LES 5yrs 6 stations.xlsx"))
    ap.add_argument("--out-dir", type=Path,
                    default=Path("met_validation/output"))
    ap.add_argument("--start", type=pd.Timestamp,
                    default=pd.Timestamp("2022-07-01 00:00"))
    ap.add_argument("--end", type=pd.Timestamp,
                    default=pd.Timestamp("2022-09-30 22:00"))
    ap.add_argument("-v", "--verbose", action="count", default=1)
    args = ap.parse_args(argv)

    level = {0: logging.WARNING, 1: logging.INFO}.get(args.verbose, logging.DEBUG)
    logging.basicConfig(
        level=level, format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    args.out_dir.mkdir(parents=True, exist_ok=True)
    files_by_coord = discover_dat_files(args.data_dir)
    _LOG.info("Discovered %d coordinate buckets", len(files_by_coord))

    all_rows: list[dict] = []
    for s in STATIONS:
        pair = _match_station(s, files_by_coord)
        if not pair:
            _LOG.error("No .dat file matched station %s @ (%s, %s)",
                       s.name, s.x_km, s.y_km)
            continue
        all_rows.extend(_evaluate_station(
            s, args.xlsx, pair, args.start, args.end, args.out_dir,
        ))

    if not all_rows:
        _LOG.error("No stations evaluated")
        return 1

    df = pd.DataFrame(all_rows)
    csv_path = args.out_dir / "stats_combined.csv"
    df.to_csv(csv_path, index=False)
    _LOG.info("Wrote %s (%d rows)", csv_path, len(df))

    _write_report(df, args.out_dir, args.start, args.end)
    return 0


if __name__ == "__main__":   # pragma: no cover
    sys.exit(main())
